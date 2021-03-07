#coding: UTF-8

import os
import sys
import itertools
import configparser
import openpyxl as xl

def GetGradOnce(sheet, ysta, rows, row_ofst, xsta, cols, col_ofst):
  flg_dict = {}
  for y, x in itertools.product(range(ysta+row_ofst+1, ysta+rows+row_ofst+1), range(xsta+col_ofst+1, xsta+cols+col_ofst+1)): # +1はエクセルが1以上でないといけないため
    xidx = sheet.cell(1, x).value
    yidx = sheet.cell(y, 1).value
    tmp  = sheet.cell(y, x).value

    if (tmp != None):
      if (tmp in flg_dict):
        flg_dict[tmp].append([xidx, yidx])
      else:
        flg_dict[tmp] = [[xidx, yidx]]

  # print(flg_dict)

  return flg_dict

def GetGradCoeff(wb):
  # 設定ファイル開く
  config = configparser.ConfigParser()
  config.read("config.ini", encoding="utf-8")
  rows     = int(config["common"]["filter_rows"])
  cols     = int(config["common"]["filter_cols"])
  row_ofst = int(config["common"]["row_ofst"])
  col_ofst = int(config["common"]["col_ofst"])
  row_skip = int(config["common"]["row_skip"])
  col_skip = int(config["common"]["col_skip"])
  row_size = rows + row_ofst + row_skip
  col_size = cols + col_ofst + col_skip

  all_filter = {}
  for i in range(len(wb.worksheets)):
    sheet = wb.worksheets[i]
    # sheetの最大サイズ取得
    max_row = sheet.max_row
    max_col = sheet.max_column

    for ysta, xsta in itertools.product(range(0, max_row, row_size), range(0, max_col, col_size)):
      # 関数名生成 (辞書のkeyとして使用)
      sub_func_name = ("_" + sheet.cell(ysta+1, xsta+1).value) if (sheet.cell(ysta+1, xsta+1).value != None) else ""
      func_name = sheet.title + sub_func_name
      all_filter[func_name] = GetGradOnce(sheet, ysta, rows, row_ofst, xsta, cols, col_ofst)
  
  return all_filter

def GetGradFilter(filepath):
  # Excel開く
  wb = xl.load_workbook(filepath)

  # ファイル名取得
  filename = os.path.splitext(os.path.basename(filepath))[0]

  # フィルタ係数取得
  all_filter = GetGradCoeff(wb)

  return filename, all_filter # filename, {funcname: [xidx, yidx]}

if __name__ == "__main__":
  filepath = sys.argv[1]
  all_filter = GetGradFilter(filepath)
  print(all_filter)